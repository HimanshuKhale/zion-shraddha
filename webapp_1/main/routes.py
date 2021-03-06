

from webapp_1.main.forms import UploadFileForm
import os
import secrets
from PIL import Image
from flask import render_template, url_for, flash, redirect, request, abort, Blueprint
from webapp_1 import db, bcrypt
from flask import current_app
from webapp_1.main.forms import UpdateAccountForm, UploadFileForm, DisaposalInstructionForm
from webapp_1.models import User, Post
from flask_login import login_user, current_user, logout_user, login_required
from werkzeug.utils import secure_filename
import re
import requests
import pdfplumber
import pandas as pd 
import openpyxl
from webapp_1.shippingbill import data_enter_sb
from webapp_1.functions import disposal_instructions
main = Blueprint('main', __name__)


@main.route("/")
@main.route("/home")
def home():
    page = request.args.get('page', 1, type=int)
    posts = Post.query.order_by(Post.post_date.desc()).paginate(page=page, per_page=5)
    return render_template('home.html', posts=posts)


@main.route("/about")
def about():
    return render_template('about.html', title='About')


@main.route('/IREX', methods = ['GET', 'POST'])
@login_required
def upload_irex():
    form = UploadFileForm()
    if form.validate_on_submit():
        #save_pdf(form.file.data)
        
        wb = openpyxl.load_workbook("C:/Users/HP/Desktop/Web_App_1/webapp/webapp_1/ZION.xlsx")
        ws_irex = wb["IREX"]
        ws_irex_pci = wb["IREX-PCI"]
        for cell in ws_irex['A']:
            if cell.value is None:
                row_no = str(cell.row)
                break
        tab = cell.row+1
        fill_cell = str("A")+str(cell.row+1)
        ws_irex[str(fill_cell)] = "S_"+ str(cell.row)

        def data_enter(col, var):  
            for cell in ws_irex['A']:
                if cell.value is None:   
                    row_no = str(cell.row)
                    break
            cell_spec = str(col) + str(cell.row)
            ws_irex[str(cell_spec)] = str(var)
            wb.save("ZION.xlsx")

        def data_enter_pci(col, var):
            for cell in ws_irex_pci['A']:
                if cell.value is None:
                    row_no = str(cell.row)
                    break
            tab = cell.row+1
            fillcell = str("A")+str(cell.row+1)
            ws_irex_pci[str(fillcell)] = "S_"+ str(cell.row)


            for cell in ws_irex_pci[col]:
                if cell.value is None:
                    row_no = str(cell.row)
                    break
            cell_spec = str(col) + str(cell.row)
            ws_irex_pci[str(cell_spec)] = str(var)
            wb.save("ZION.xlsx")


        filename = secure_filename(form.file.data.filename)
        form.file.data.save('C:/Users/HP/Desktop/Web_App_1/webapp/webapp_1/static/pdf_files/'+ filename)
        fullpath = os.path.join("C:/Users/HP/Desktop/Web_App_1/webapp/webapp_1/static/pdf_files/"+ filename)
        #open the advice copy
        with pdfplumber.open(fullpath) as irex_pdf:
            page = irex_pdf.pages[0]
            text = page.extract_text()

        
        #get irex number
        irex_ref = re.compile(r'0500IREX\d{8}')
        matches = irex_ref.finditer(text)

        for match in matches:
            irex_no = match.group(0)
            print(match.group(0))

        #get PCI number
        pci_ref = re.compile(r"^3174PCI.*")
        matches = pci_ref.finditer(text)

        for match in matches:
            pci_no = match.groups()
            print(match.group(0))
            #print(pci_no)
            
        #get PCI Number - longer approach
        pci_refs = re.compile(r'3174PCI\d+\s\d+\s\w+.\w+.\w+.\w+')
        matches = pci_refs.finditer(text)

        for match in matches:
            pci_details = match.group(0)
            #print(match.group(0))
        # print(match)
            print(pci_details)
            

        #get remitter name
        remitter = re.compile(r'REMITTERNAME\s\w+')
        matches = remitter.finditer(text)

        for match in matches:
            remitter_name = match.group(0)
            print(match.group(0))
            remitter = remitter_name.replace("REMITTERNAME ", "")

        #get value date
        valuedate = re.compile(r'VALUEDATE\s\d{2}(/)\d{2}(/)\d{4}')
        matches = valuedate.finditer(text)

        for match in matches:
            date = match.group(0)
            print(match.group(0))
            value_date = date.replace("VALUEDATE ", "")

        amount_inr = re.compile(r'USD\s\d+.+')
        matches = amount_inr.finditer(text)

        for match in matches:
            amt_1=match.group(0)
            print(match.group(0))
            fc, amt_fc, rate, inr, amt_inr  = amt_1.split(" ")
            print(amt_fc)


        pci_mat = re.findall('3174PCI\d+\s\d+\s\w+.\w+.\w+.\w+', text)
        print(pci_mat)
        index = range(1,7)
        length = len(pci_mat)
        iter_range = range(0, length)
        for i in iter_range:
            pc_elem = str(pci_mat[i])
            print(pc_elem)
            
            pc_no, pc_acc, remb_amt = pc_elem.split(" ")
            print(pc_no, remb_amt)
            data_enter_pci("B", irex_no)
            data_enter_pci("C", value_date)
            data_enter_pci("D", amt_inr)
            data_enter_pci("E", pc_no)
            data_enter_pci("F", remb_amt)

        

        
        
        col_dict = {1:"B", 2:"C", 3:"D", 4:"E", 5:"F", 6:"G"}
        var_dict = {1:irex_no, 2:value_date, 3:amt_fc, 4:rate, 5:amt_inr, 6:remitter}
       

        dict_range = range(1-7)
        for i in index:
            data_enter(col_dict.get(i), var_dict.get(i))

        wb.save("C:/Users/HP/Desktop/Web_App_1/webapp/webapp_1/ZION.xlsx")
        print(irex_no, value_date, amt_fc, rate, amt_inr, remitter)


        flash('Advice copy uploaded successfully' , 'success')
        print(fullpath)
        return redirect(url_for('main.home'))
    

    return render_template('upload_file.html', form=form)





@main.route('/PCI', methods = ['GET', 'POST'])
@login_required
def upload_pci():
    form = UploadFileForm()
    if form.validate_on_submit():
        #save_pdf(form.file.data)
        wb = openpyxl.load_workbook("C:/Users/HP/Desktop/Web_App_1/webapp/webapp_1/ZION.xlsx")
        ws_irex = wb['IREX']
        ws_irex_pci = wb['IREX-PCI']
        ws_epc = wb['PCS']



        ### Worksheet EPC and data entry function

        for cell in ws_epc['A']:
            if cell.value is None:
                row_no = str(cell.row)
                break
        tab = cell.row+1
        fillcell = str("A")+str(cell.row+1)
        ws_epc[str(fillcell)] = "S_"+ str(cell.row)

        def data_enter_irex(col, var):
            for cell in ws_epc['A']:
                if cell.value is None:   
                    row_no = str(cell.row)
                    break
            cell_spec = str(col) + str(cell.row)
            ws_epc[str(cell_spec)] = str(var)
            wb.save("ZION.xlsx")

        filename = secure_filename(form.file.data.filename)
        form.file.data.save('C:/Users/HP/Desktop/Web_App_1/webapp/webapp_1/static/pdf_files/'+ filename)
        fullpath = os.path.join("C:/Users/HP/Desktop/Web_App_1/webapp/webapp_1/static/pdf_files/"+ filename)
        #open the advice copy
        
        
        with pdfplumber.open(fullpath) as pdf:
            page = pdf.pages[0]
            text = page.extract_text()

        #print(text)

        pattern = re.compile(r'3174PCI\d{9}')
        matches = pattern.finditer(text)

        for match in matches:
            EPC_no = match.group(0)
            #print(match.group(0))

        disb_date = re.compile(r'DisbursementDate\s\d+.+')
        matches = disb_date.finditer(text)

        for match in matches:
            date_disb = match.group(0)
            tag, date = date_disb.split(" ")
            
            #print(match.group(0))

        po = re.compile(r'SI/\w+.+')
        matches = po.finditer(text)

        for match in matches:
            agreement_no = match.group(0)
            #print(match.group(0))

        buyer = re.compile(r'OverseasBuyerName\s\w+.+')
        matches = buyer.finditer(text)

        for match in matches:
            buyer_det = match.group(0)
            tag, buyer = buyer_det.split(" ")
            #print(match.group(0))

        EPC_details = re.compile(r'\d{2}.\d{2}.\d{4}\s\d{2}.\d{2}.\d{4}\s\w+.+')
        matches = EPC_details.finditer(text)

        for match in matches:
            epc_det = match.group(0)
            disbursedate, DueDate, Tenure, Amount, Interest = epc_det.split(" ")
            #print(match.group(0))


        epc_dict = {1:"B", 2:"C", 3:"D", 4:"E", 5:"F", 6:"G", 7:"H", 8:"I"}
        epc_var_dict = {1:EPC_no, 2:date, 3:agreement_no, 4:DueDate, 5:Tenure, 6:buyer, 7:Amount, 8:Interest}


        for i in range(1,9):
            data_enter_irex(epc_dict.get(i), epc_var_dict.get(i))

        wb.save("ZION.xlsx")


        flash('Advice copy uploaded successfully' , 'success')
        print(fullpath)
        return redirect(url_for('main.home'))



    return render_template('upload_file.html', form=form)

@main.route("/upload-docs", methods=['GET', 'POST'])
@login_required
def upload_docs():
    #form = UpdateAccountForm()
    return render_template('Upload_Page.html')

@main.route("/Upload-Shipping_Bill", methods=['GET', 'POST'])
@login_required
def upload_sb():
    form = UploadFileForm()
    if form.validate_on_submit():
        #save_pdf(form.file.data)
        
       

        
        
        wb = openpyxl.load_workbook("C:/Users/HP/Desktop/Web_App_1/webapp/webapp_1/ZION.xlsx")
        ws_irex = wb['IREX']
        ws_irex_pci = wb['IREX-PCI']
        ws_epc = wb['PCS']
        ws_shipping = wb['SB']

        for cell in ws_shipping['A']:

            if cell.value is None:
                row_no = str(cell.row)
                break
            fillcell = str("A") + str(cell.row+1)
            ws_shipping[str(fillcell)] = "S_" + str(cell.row)
            
        

        def data_sb(col, var):
            
            for cell in ws_shipping['A']:
                if cell.value is None:
                    row_no=str(cell.row)
                    break
            cell_spec = str(col) + str(cell.row)
            ws_shipping[str(cell_spec)] = str(var)
            wb.save("ZION.xlsx")

        filename = secure_filename(form.file.data.filename)
        form.file.data.save('C:/Users/HP/Desktop/Web_App_1/webapp/webapp_1/static/pdf_files/'+ filename)
        fullpath = os.path.join("C:/Users/HP/Desktop/Web_App_1/webapp/webapp_1/static/pdf_files/"+ filename)

        with pdfplumber.open(fullpath) as sb_pdf:
            page1 = sb_pdf.pages[0]
            page2 = sb_pdf.pages[1]
            text1 = page1.extract_text()
            text2 = page2.extract_text()
    

        #Notify Party
        notify = re.compile(r'SHRADDHA IMPEX.+')
        matches = notify.finditer(text2)

        for match in matches:
            notify_party = match.group(0)
            print(match.group(0))
            notifyparty1 = notify_party.replace("SHRADDHA IMPEX ", '')
            notifyparty = notifyparty1.replace("SHRADDHA IMPEX ", '')
            print(notifyparty)
            data_sb('G', notifyparty)

        #Invoice No. 
        invoice_details = re.compile(r' SI.+USD')
        matches = invoice_details.finditer(text1)

        for match in matches:
            invoice = match.group(0)
            tag, invoiceno, giv, tag2 = invoice.split(" ")
            print(match.group(0))
            data_sb('C', invoiceno)
            print(invoiceno, "\n", giv)

        #Package No, Gross Wt, Shipping Bill No
        grs_wt = re.compile(r'PKG\s.+')
        matches = grs_wt.finditer(text1)

        for match in matches:
            gross_wt = match.group(0)
            pkg, pkg_no, gwt, mt, grosswt, SB_No = gross_wt.split(" ")
            print(match.group(0))
            data_sb('K', pkg_no)
            data_sb('J', grosswt)
            data_sb('B', SB_No)
            print(pkg_no, "\n", grosswt, "\n", SB_No)


        #Leo Date
        leodt = re.compile(r'LEO Date.+')
        matches = leodt.finditer(text1)
        for match in matches:
            leodate = match.group(0)
            tag3, tag4, LEO_Date = leodate.split(" ")
            print(match.group(0))
            data_sb('H', LEO_Date)
            print(LEO_Date)

        #Consignee
        consignee = re.compile(r'SHRADDHA.+')
        matches = consignee.finditer(text1)
        for match in matches:
            print(match.group(0))
            cnsgni = match.group(0)
            consignee_name = cnsgni.replace("SHRADDHA IMPEX ", '')
            print(consignee_name)
            #data_sb('G', consignee_name)

        #Invoice Date
        invoicedate = re.compile(r'SI.+\s\d{2}.\d{2}.\d{4}')
        matches = invoicedate.finditer(text2)
        for match in matches:
            print(match.group(0))
            invdttag = match.group(0)
            tag5, invoice_date = invdttag.split(" ")
            data_sb('D', invoice_date)

            
        print(tag, invoice_date)
        #Manufacturer
        mfger = re.compile(r"MAEQ QUOTA\s.+")
        matches = mfger.finditer(text2)
        for match in matches:
            print(match.group(0))
        #HS Code
        hscode = re.compile(r'1\s\d{8}')
        matches = hscode.finditer(text2)
        for match in matches:
            print(match.group(0))
            hs = match.group(0)
            tag5, hs_code = hs.split(" ")
            print(hs_code)
            data_sb('I', hs_code)
        #Mfg invoice no. 
        splrinv = re.compile(r"s(/)2021(-)22(/)\d{4}")
        matches = splrinv.finditer(text2)
        for match in matches:
            print(match.group(0))
        #Destination 
        discharge_country = re.compile(r'COUNTRY OF DISCHARGE\s.+')
        matches = discharge_country.finditer(text1)
        for match in matches:
            print(match.group(0))
            country = match.group(0)
            country_of_discharge = country.replace("COUNTRY OF DISCHARGE ", "")
            print(country_of_discharge)
            
        #Dest Port
        destport = re.compile(r'PORT OF FINAL DESTINATION\s.+')
        matches = destport.finditer(text1)
        for match in matches:
            print(match.group(0))
            portofdest = match.group(0)
            destination_port = portofdest.replace("PORT OF FINAL DESTINATION ", "")
            print(destination_port)


        sb_dict = {1:"B", 2:"C", 3:"D", 4:"E", 5:"F", 6:"G", 7:"H", 8:"I", 9:"J", 10:"K"}
        sb_var_dict = {1:SB_No, 2:invoiceno, 3:invoice_date, 4:giv, 5:consignee_name, 6:notifyparty, 7:LEO_Date, 8:hs_code, 9:grosswt, 10:pkg_no}

        index = range(1, 11)

        for i in index:
            data_sb(sb_dict.get(i) , sb_var_dict.get(i))
            wb.save("ZION.xlsx")

        data_enter_sb("L", destination_port)
            

        print("Data Entry Successful")

        
        flash("The shipping bill data has been updated", "success")
        return redirect(url_for('main.home'))
    return render_template('upload_file.html', form=form)

@main.route("/disposal_instruction", methods = ['GET', 'POST'])
@login_required
def disposal():
    form = DisaposalInstructionForm()
    if form.validate_on_submit():
        party = form.party.data
        amount_ = form.amount.data

        disposal_instructions(party, amount_)

  
    
    return render_template('disposal.html', form=form)


