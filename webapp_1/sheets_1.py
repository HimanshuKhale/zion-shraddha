import openpyxl
wb = openpyxl.load_workbook("SI92.xlsx")
ws = wb['Sheet1']
ws
for cell in ws["A"]:
    if cell.value is None:
        print(cell.row)
        break

print(cell.row + 1)
tab = cell.row +1
fillcell = str("A") + str(cell.row+1)
print(fillcell)
ws[str(fillcell)] = "S_" + str(cell.row)

IREX = input("irex = ?")

def data_enter(col, var):
    for cell in ws['A']:
        if cell.value is None:
            row_no = str(cell.row)
            break
    cell_spec = str(col) + str(cell.row+1)
    ws[str(cell_spec)] = str(var)

data_enter("C", IREX)

wb.save("SI92.xlsx")    